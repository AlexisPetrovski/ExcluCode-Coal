import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import re
import io

# ──────────────────────────────────────────────────────────────────────────────
# Helper: robust conversion to float (handles EU & US formats)
# ──────────────────────────────────────────────────────────────────────────────
# 🔹 1. Removes spaces, Detects comma vs dot for decimals, Converts safely to float, Returns 0.0 on error 🔹
def to_float(val):
    s = str(val).strip().replace(" ", "")
    if s == "" or s.lower() in ("nan", "none"):
        return 0.0
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        s = s.replace(",", ".") if len(parts) == 2 and len(parts[1]) in (1, 2) else s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# 🔹 Ensures that repeated column names like "Company", "Company" are renamed to "Company", "Company_1". If it's the first appearance of the column → use the name as-is ("Company"). If it's a duplicate → add a numeric suffix, like "Company_1", "Company_2" 🔹
def make_columns_unique(df):
    seen, new_cols = {}, []
    for c in df.columns:
        seen[c] = seen.get(c, 0) + 1
        new_cols.append(c if seen[c] == 1 else f"{c}_{seen[c]-1}")
    df.columns = new_cols
    return df

# 🔹 It renames columns in a table (a DataFrame) by looking for similar names, even if they're not exactly the same. 🔹
def fuzzy_rename_columns(df, rename_map):
    used = set()
    for final_name, pats in rename_map.items():
        for col in df.columns:     # 🔹 Iterates through each column name (col) in the DataFrame. 🔹
            if col in used:        # 🔹 If this column has already been renamed, skip it and go to the next one. 🔹
                continue
            # 🔹 Avoid renaming "Parent Company" to "Company" — it's considered a special, unneeded field that should be ignored. 🔹
            if final_name == "Company" and col.strip().lower() == "parent company":
                continue
            if any(p.lower().strip() in col.lower() for p in pats): # 🔹 pats is string that comes from dictionary, from rename_map dictionary. While p is individual string in this pats list. Each p is compared against the actual column name (col) in the line: 🔹
                df.rename(columns={col: final_name}, inplace=True)
                used.add(col)
                break
    return df

# 🔹 Normatlization of string 🔹
def normalize_key(s):
    s = str(s).lower()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\b(ltd|limited|sa|plc|inc|corp|co)\b", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

# 🔹 loaders and data preperation that headers would be in proper place and remove double rows 🔹
def load_spglobal(file, sheet_name="Sheet1"):
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        data = list(ws.values)
        full_df = pd.DataFrame(data)
        if len(full_df) < 6:
            raise ValueError("SPGlobal file does not have enough rows.")

        row5 = full_df.iloc[4].fillna("")
        row6 = full_df.iloc[5].fillna("")
        final_cols = []
        for a, b in zip(row5, row6):
            top = str(a).strip(); bot = str(b).strip()    # 🔹 Strips whitespace and ensures both a and b are strings. 🔹
            col = top if top else ""                      # 🔹 If the top label exists, use it as the base name. If not, start with an empty string. 🔹
            if bot and bot.lower() not in col.lower():    # 🔹 If there's a bottom label (bot) and it's not already in the col name: Append it (with a space in between). Ensures no repeated label parts like "Revenue Revenue" 🔹
                col = f"{col} {bot}".strip()
            final_cols.append(col)

        sp_df = full_df.iloc[6:].reset_index(drop=True)   # 🔹 Selects all rows starting from row 7 (index 6) to the end of the DataFrame .reset_index(drop=True): Resets the row index to start at 0 and drops the old index. 🔹
        sp_df.columns = final_cols                        # 🔹 Assigns the cleaned and merged column names (from earlier using zip(row5, row6)) to the DataFrame. 🔹
        sp_df = make_columns_unique(sp_df)

        # 🔹 Standardized names for integrity🔹
        rename_map_sp = {
            "SP_ENTITY_NAME": ["sp entity name", "entity name"],
            "SP_ENTITY_ID": ["sp entity id", "entity id"],
            "SP_COMPANY_ID": ["sp company id", "company id"],
            "SP_ISIN": ["sp isin"],
            "SP_LEI": ["sp lei"],
            "Generation (Thermal Coal)": ["generation (thermal coal)"],
            "Thermal Coal Mining": ["thermal coal mining"],
            "Coal Share of Revenue": ["coal share of revenue"],
            "Coal Share of Power Production": ["coal share of power production"],
            "Installed Coal Power Capacity (MW)": ["installed coal power capacity"],
            "Coal Industry Sector": ["industry sector"],
            ">10MT / >5GW": [">10mt", ">5gw"],
            "expansion": ["expansion"],
        }
        sp_df = fuzzy_rename_columns(sp_df, rename_map_sp).astype(object)

        # 🔹 This is a list of key numeric columns in the file that need to be properly converted to float numbers (like percentages or MW values). 🔹
        for col in [
            "Thermal Coal Mining", "Generation (Thermal Coal)",
            "Coal Share of Revenue", "Coal Share of Power Production",
            "Installed Coal Power Capacity (MW)", "Annual Coal Production (in million metric tons)"
        ]:
            if col in sp_df:
                sp_df[col] = sp_df[col].apply(to_float)

        return sp_df
    except Exception as e:
        st.error(f"Error loading SPGlobal: {e}")
        return pd.DataFrame()

# 🔹 The load_urgewald function reads the Urgewald Excel file🔹
def load_urgewald(file, sheet_name=None):
    try:
        wb = openpyxl.load_workbook(file, data_only=True)

        # ✅ FIX: auto-detect GCEL sheet if not provided
        if sheet_name is None or sheet_name not in wb.sheetnames:
            for s in wb.sheetnames:
                if "gcel" in s.lower():
                    sheet_name = s
                    break

        # 🚨 fallback if nothing found
        if sheet_name not in wb.sheetnames:
            st.error(f"❌ No GCEL sheet found. Available sheets: {wb.sheetnames}")
            return pd.DataFrame()

        ws = wb[sheet_name]

        # 🔍 DEBUG
        st.write(f"Using Urgewald sheet: {sheet_name}")

        data = list(ws.values)
        full_df = pd.DataFrame(data)
        if full_df.empty:
            raise ValueError("Urgewald file is empty.")
# 🔹 Removes any "Parent Company" columns, cleans up the headers, fixes duplicate column names, and converts key numbers (like revenue %) into proper numeric values. It prepares the data so it's ready to be used in filtering or analysis.🔹
        header = full_df.iloc[0].fillna("")
        keep = header.str.strip().str.lower() != "parent company"
        ur_df = full_df.iloc[1:].reset_index(drop=True).loc[:, keep]
        ur_df.columns = [c for c in header if str(c).strip().lower() != "parent company"]
        ur_df = make_columns_unique(ur_df)

        # ✅ FIX: normalize column names early
        ur_df.columns = [str(c).strip() for c in ur_df.columns]
       
        # 🔹 Standardized names for integrity🔹
        rename_map_ur = {
            "Company": ["company", "issuer name"],
            "ISIN equity": ["isin equity", "isin(eq)", "isin eq"],
            "LEI": ["lei"],
            "BB Ticker": ["bb ticker"],
            "Coal Industry Sector": ["industry sector"],
            ">10MT / >5GW": [">10mt", ">5gw"],
            "expansion": ["expansion"],
            "Coal Share of Power Production": ["coal share of power production"],
            "Coal Share of Revenue": ["coal share of revenue"],
            "Installed Coal Power Capacity (MW)": ["installed coal power capacity"],
            "Generation (Thermal Coal)": ["generation (thermal coal)"],
            "Thermal Coal Mining": ["thermal coal mining"],
        }
        ur_df = fuzzy_rename_columns(ur_df, rename_map_ur).astype(object)

        # 🔍 DEBUG: check columns after renaming
        st.write("Urgewald columns after rename:", ur_df.columns.tolist())

        # ✅ FIX: ensure 'Company' column exists
        if "Company" not in ur_df.columns:
            for col in ur_df.columns:
                if "company" in str(col).lower():
                    ur_df.rename(columns={col: "Company"}, inplace=True)
                    break

            # 🚨 HARD FAIL if still missing
            if "Company" not in ur_df.columns:
                st.error("❌ 'Company' column not detected in Urgewald file")

        # 🔹 This is a list of key numeric columns in the file that need to be properly converted to float numbers (like percentages or MW values). 🔹
        for col in [
            "Thermal Coal Mining", "Generation (Thermal Coal)",
            "Coal Share of Revenue", "Coal Share of Power Production",
            "Installed Coal Power Capacity (MW)", "Annual Coal Production (in million metric tons)"
        ]:
            if col in ur_df:
                ur_df[col] = ur_df[col].apply(to_float)

        return ur_df
    except Exception as e:
        st.error(f"Error loading Urgewald: {e}")
        return pd.DataFrame()


# 🔹 Merge 🔹
def merge_ur_into_sp_opt(sp_df, ur_df):
    sp = sp_df.copy().astype(object)
    ur = ur_df.copy().astype(object)

    # 🔹 Creating normalized versions of key identity columns from the SPGlobal dataset 🔹
    if "SP_ISIN" not in sp.columns:
        sp["SP_ISIN"] = ""   # create an empty column if it doesn't exist
        
    sp["norm_isin"] = sp["SP_ISIN"].astype(str).apply(normalize_key)

    if "SP_LEI" not in sp.columns:
        sp["SP_LEI"] = ""
    
    sp["norm_lei"] = sp["SP_LEI"].astype(str).apply(normalize_key)

    if "SP_ENTITY_NAME" not in sp.columns:
        sp["SP_ENTITY_NAME"] = ""
    
    sp["norm_name"] = sp["SP_ENTITY_NAME"].astype(str).apply(normalize_key)

    # 🔹 Creating normalized versions of key identity columns from the Urgewald dataset and searches for identities in the raw 🔹
    for col in ["ISIN equity", "LEI", "Company"]:
        if col not in ur:
            ur[col] = ""
    ur["norm_isin"] = ur["ISIN equity"].astype(str).apply(normalize_key)
    ur["norm_lei"] = ur["LEI"].astype(str).apply(normalize_key)
    ur["norm_company"] = ur["Company"].astype(str).apply(normalize_key)

    # 🔹 Creating a dictionary with the identifier and index i. These maps make it quick and easy to match a company from the Urgewald file to a row in the SP file based on a shared ID like ISIN, LEI, or name. 🔹
    isin_map = {k: i for i, k in enumerate(sp["norm_isin"]) if k}
    lei_map = {k: i for i, k in enumerate(sp["norm_lei"]) if k}
    name_map = {k: i for i, k in enumerate(sp["norm_name"]) if k}

    # 🔹 ur_not is empty list for not matched companies and look for matching. If a match is found, the target becomes the index of the SPGlobal row to merge into. 🔹
    ur_not = []
    for _, r in ur.iterrows():
        target = None
        if r["norm_isin"] in isin_map:
            target = isin_map[r["norm_isin"]]
        elif r["norm_lei"] in lei_map:
            target = lei_map[r["norm_lei"]]
        elif r["norm_company"] in name_map:
            target = name_map[r["norm_company"]]

        # 🔹 This lets you enrich SPGlobal data with missing info from Urgewald — only when they match by ISIN, LEI, or name. 🔹
        if target is not None:
            for c, v in r.items():
                if c.startswith("norm_"):    # 🔹 These columns like "norm_isin", "norm_lei" etc. were only used for matching — they are not real data. So we skip them. We don't want to merge them into the final DataFrame. 🔹
                    continue
                if c not in sp or pd.isna(sp.at[target, c]) or str(sp.at[target, c]).strip() == "":  # 🔹 Filling other missing identifiers if found in urgewald 🔹
                    sp.at[target, c] = v
            sp.at[target, "Merged"] = True
        else:
            ur_not.append(r)
    # 🔹 This code finalizes the merge by clearly marking which rows were merged and which were not.🔹 
    if "Merged" not in sp.columns:
        sp["Merged"] = False
    else:
        sp["Merged"] = sp["Merged"].fillna(False)
   ur_only = pd.DataFrame(ur_not)

        # ✅ FIX: ensure Merged column always exists
            if "Merged" not in ur_only.columns:
                ur_only["Merged"] = False
                    else:
                ur_only["Merged"] = ur_only["Merged"].fillna(False)



    for c in [c for c in sp.columns if c.startswith("norm_")]:
        sp.drop(columns=c, inplace=True, errors="ignore")    # 🔹  Removes any temporary norm_* columns (like norm_isin, norm_lei, norm_name) that were used for matching. 🔹
    for c in [c for c in ur_only.columns if c.startswith("norm_")]:
        ur_only.drop(columns=c, inplace=True, errors="ignore") # 🔹  Removes any temporary norm_* columns (like norm_isin, norm_lei, norm_name) that were used for matching. 🔹

    return sp, ur_only


# 🔹 It allows  code to dynamically switch between > and ≥ comparisons — based on a checkbox or user input in the Streamlit UI.🔹
def test(val, thr, ge):
    """Return True if value triggers rule ( > or ≥ )."""
    return val >= thr if ge else val > thr
    
def op(ge: bool) -> str:
    return "≥" if ge else ">"

# 🔹🔹🔹 Compute_exclusion 🔹🔹🔹   
def compute_exclusion(row, **params):
    reasons = []

# 🔹 Dataset with checking whether values are percentages 🔹 
    sp_min = row.get("Thermal Coal Mining", 0.0)
    sp_pow = row.get("Generation (Thermal Coal)", 0.0)

    ur_rev_pct = row.get("Coal Share of Revenue", 0.0)
    ur_rev_pct = ur_rev_pct if ur_rev_pct > 1 else ur_rev_pct * 100
    ur_pp_pct = row.get("Coal Share of Power Production", 0.0)
    ur_pp_pct = ur_pp_pct if ur_pp_pct > 1 else ur_pp_pct * 100

    # 🔹 This section extracts and prepares key values from each row of the data to help check exclusion rules:🔹
    prod_str = str(row.get(">10MT / >5GW", "")).lower()
    cap = row.get("Installed Coal Power Capacity (MW)", 0.0)
    expansion = str(row.get("expansion", "")).lower()

    has_sp = bool(str(row.get("SP_ENTITY_NAME", "")).strip())
    # ✅ FIX: more robust Urgewald detection
    has_ur = any([
        str(row.get("Company", "")).strip(),
        str(row.get("ISIN equity", "")).strip(),
        str(row.get("LEI", "")).strip()
        ])



    # 🔹 sectors 🔹
    sector_raw = str(row.get("Coal Industry Sector", "")).lower()
    mining_kw = ("mining", "extraction", "producer")
    power_kw = ("power", "generation", "utility", "electric")
    tokens = [p.strip() for p in re.split(r"[;,/]|(?:\s*\n\s*)", sector_raw) if p.strip()]
    mining_parts = [p for p in tokens if any(k in p for k in mining_kw)]
    power_parts = [p for p in tokens if any(k in p for k in power_kw)]
    other_parts = [p for p in tokens if p not in mining_parts + power_parts]

     # 🔹 Ensures that company only in specific sector🔹
    is_mining_only = bool(mining_parts) and not power_parts and not other_parts
    is_power_only = bool(power_parts) and not mining_parts and not other_parts
    is_mixed = bool(mining_parts) and bool(power_parts) and not other_parts

    # ✅ FIX: fallback if sector parsing fails
    if has_ur and not (is_mining_only or is_power_only or is_mixed):
        is_mixed = True

    # 🔹 This code block adds exclusion reasons based on general (non-revenue) filters if certain conditions are met. 🔹
    if params["exclude_mt"] and "10mt" in prod_str:
        reasons.append(">10 MT indicator")

    if params["exclude_capacity"] and test(cap, params["capacity_threshold"], params["capacity_ge"]):
        reasons.append(
            f"Installed capacity {cap:.0f} MW {op(params['capacity_ge'])} {params['capacity_threshold']:.0f} MW"
        )

    if params["exclude_power_prod"] and test(ur_pp_pct, params["power_prod_threshold"], params["power_prod_ge"]):
        reasons.append(
            f"Coal power production {ur_pp_pct:.2f}% {op(params['power_prod_ge'])} {params['power_prod_threshold']}%"
        )

    # 🔹 S&P rules with exclusion reasons output 🔹
    if has_sp:
        if params["sp_mining_checkbox"] and test(sp_min, params["sp_mining_threshold"], params["sp_mining_ge"]):
            reasons.append(
                f"SP mining revenue {sp_min:.2f}% {op(params['sp_mining_ge'])} {params['sp_mining_threshold']}%"
            )
        if params["sp_power_checkbox"] and test(sp_pow, params["sp_power_threshold"], params["sp_power_ge"]):
            reasons.append(
                f"SP power revenue {sp_pow:.2f}% {op(params['sp_power_ge'])} {params['sp_power_threshold']}%"
            )
        if params["sp_level2_checkbox"]:
            combo = sp_min + sp_pow
            if test(combo, params["sp_level2_threshold"], params["sp_level2_ge"]):
                reasons.append(
                    f"SP level-2 combined {combo:.2f}% {op(params['sp_level2_ge'])} {params['sp_level2_threshold']}%"
                )

    # 🔹 This code block checks Urgewald revenue-based exclusion rules — and adds detailed reasons for exclusion based on the company's sector type (mining, power, or mixed) and revenue percentage.🔹
    if has_ur:
        if is_mining_only and params["ur_mining_checkbox"] and test(ur_rev_pct, params["ur_mining_threshold"], params["ur_mining_ge"]):
            reasons.append(
                f"UR mining revenue {ur_rev_pct:.2f}% {op(params['ur_mining_ge'])} {params['ur_mining_threshold']}%"
            )
        if is_power_only and params["ur_power_checkbox"] and test(ur_rev_pct, params["ur_power_threshold"], params["ur_power_ge"]):
            reasons.append(
                f"UR power revenue {ur_rev_pct:.2f}% {op(params['ur_power_ge'])} {params['ur_power_threshold']}%"
            )
        if is_mixed and params["ur_mixed_checkbox"] and test(ur_rev_pct, params["ur_mixed_threshold"], params["ur_mixed_ge"]):
            reasons.append(
                f"UR mixed revenue {ur_rev_pct:.2f}% {op(params['ur_mixed_ge'])} {params['ur_mixed_threshold']}%"
            )
        if params["ur_level2_checkbox"] and test(ur_rev_pct, params["ur_level2_threshold"], params["ur_level2_ge"]):
            reasons.append(
                f"UR level-2 revenue {ur_rev_pct:.2f}% {op(params['ur_level2_ge'])} {params['ur_level2_threshold']}%"
            )

    # 🔹 This short block checks if a company mentions coal expansion in its description — and flags it for exclusion if any excluded keywords are found. 🔹
    for kw in params["expansion_exclude"]:
        if kw.lower() in expansion:
            reasons.append(f"Expansion matched '{kw}'")
            break
   
    # 🔹 Output with all exclusion reasons combined 🔹
    return pd.Series([bool(reasons), "; ".join(reasons)], index=["Excluded", "Exclusion Reasons"])
    
# 🔹 Streamlit UI (added individual ≥ toggles) 🔹
def main():
    st.set_page_config(page_title="Coal Exclusion Filter", layout="wide")
    st.title("Coal Exclusion Filter")

    # 🔹 file inputs 🔹
    st.sidebar.header("File & Sheet Settings")
    sp_sheet = st.sidebar.text_input("SPGlobal Sheet Name", "Sheet1")
    ur_sheet = st.sidebar.text_input("Urgewald Sheet Name (optional)", "")
    sp_file = st.sidebar.file_uploader("Upload SPGlobal Excel file", type=["xlsx"])
    ur_file = st.sidebar.file_uploader("Upload Urgewald Excel file", type=["xlsx"])
    st.sidebar.markdown("---")

    # 🔹 helper: numeric + ≥ 🔹 
    def num_ge(label, default, key):
        c1, c2 = st.columns([3, 1])
        with c1:
            v = st.number_input(label, value=default, key=f"{key}_v")
        with c2:
            g = st.checkbox("≥", value=False, key=f"{key}_ge")
        return v, g

    # 🔹 Mining expander (unchanged block order) 🔹
    with st.sidebar.expander("Mining", True):
        ur_mining_checkbox = st.checkbox("UR: Exclude mining-only", False)
        ur_mining_threshold, ur_mining_ge = num_ge("UR Mining threshold (%)", 5.0, "ur_min")
        sp_mining_checkbox = st.checkbox("SP: Exclude mining-only", True)
        sp_mining_threshold, sp_mining_ge = num_ge("SP Mining threshold (%)", 5.0, "sp_min")
        exclude_mt = st.checkbox("Exclude >10MT", True)
        mt_threshold = st.number_input("MT threshold (informational)", value=10.0)

    # 🔹 Power expander 🔹 
    with st.sidebar.expander("Power", True):
        ur_power_checkbox = st.checkbox("UR: Exclude power-only", False)
        ur_power_threshold, ur_power_ge = num_ge("UR Power threshold (%)", 20.0, "ur_pow")
        sp_power_checkbox = st.checkbox("SP: Exclude power-only", True)
        sp_power_threshold, sp_power_ge = num_ge("SP Power threshold (%)", 20.0, "sp_pow")
        exclude_power_prod = st.checkbox("Exclude power-production %", True)
        power_prod_threshold, power_prod_ge = num_ge("Power-production threshold (%)", 20.0, "ppp")
        exclude_capacity = st.checkbox("Exclude installed capacity", True)
        capacity_threshold, capacity_ge = num_ge("Capacity threshold (MW)", 10000.0, "cap")

    # 🔹 UR Mixed Level-1 🔹
    with st.sidebar.expander("UR Mixed Level 1", False):
        ur_mixed_checkbox = st.checkbox("UR: Exclude mining & power", False)
        ur_mixed_threshold, ur_mixed_ge = num_ge("UR Mixed threshold (%)", 25.0, "ur_mix")

    # 🔹 UR Level-2 🔹
    with st.sidebar.expander("UR: mining, power and services (Level 2)", False):
        ur_level2_checkbox = st.checkbox("UR: mining, power and services", False)
        ur_level2_threshold, ur_level2_ge = num_ge("Revenue threshold (%)", 10.0, "ur_l2")

    # 🔹 SP Level-2 🔹
    with st.sidebar.expander("SP:mining and power (Level 2)", False):
        sp_level2_checkbox = st.checkbox("SP:mining and power", False)
        sp_level2_threshold, sp_level2_ge = num_ge("Revenue threshold (%)", 10.0, "sp_l2")

    # 🔹 expansion 🔹
    with st.sidebar.expander("Exclude expansions", False):
        expansions_possible = ["mining", "infrastructure", "power", "subsidiary of a coal developer"]
        expansion_exclude = st.multiselect("Exclude if expansion text contains", expansions_possible, [])

    st.sidebar.markdown("---")
    if not st.sidebar.button("Run"):
        st.stop()

    # 🔹 load 🔹
    if not sp_file or not ur_file:
        st.warning("Please upload both files")
        st.stop()
    sp_df = load_spglobal(sp_file, sp_sheet)
    ur_df = load_urgewald(ur_file, ur_sheet if ur_sheet.strip() else None)
    if sp_df.empty or ur_df.empty:
        st.warning("Error loading data")
        st.stop()
        
    # 🔹 This code ensures that the Merged column is correctly set for both datasets after trying to match and merge the SPGlobal and Urgewald data. 🔹
    merged_sp, ur_only = merge_ur_into_sp_opt(sp_df, ur_df)
        # 🔍 DEBUG: check merge results
            st.write("SP rows:", len(sp_df))
            st.write("UR rows:", len(ur_df))
            st.write("Merged rows:", merged_sp["Merged"].sum())
            st.write("UR only rows:", len(ur_only))
    for d in (merged_sp, ur_only):
        d["Merged"] = d.get("Merged", False).fillna(False)

    sp_merged = merged_sp[merged_sp.Merged]
    sp_only = merged_sp[~merged_sp.Merged & (
        (merged_sp["Thermal Coal Mining"] > 0) | (merged_sp["Generation (Thermal Coal)"] > 0)
    )]
    ur_unmerged = ur_only[~ur_only.Merged]

    # 🔹 This dictionary holds all the toggle values, thresholds, and comparison types (≥ or >) chosen by the user, so the filtering logic later can work properly.🔹
    params = dict(
        ur_mining_checkbox=ur_mining_checkbox, ur_mining_threshold=ur_mining_threshold, ur_mining_ge=ur_mining_ge,
        sp_mining_checkbox=sp_mining_checkbox, sp_mining_threshold=sp_mining_threshold, sp_mining_ge=sp_mining_ge,
        ur_power_checkbox=ur_power_checkbox, ur_power_threshold=ur_power_threshold, ur_power_ge=ur_power_ge,
        sp_power_checkbox=sp_power_checkbox, sp_power_threshold=sp_power_threshold, sp_power_ge=sp_power_ge,
        ur_mixed_checkbox=ur_mixed_checkbox, ur_mixed_threshold=ur_mixed_threshold, ur_mixed_ge=ur_mixed_ge,
        ur_level2_checkbox=ur_level2_checkbox, ur_level2_threshold=ur_level2_threshold, ur_level2_ge=ur_level2_ge,
        sp_level2_checkbox=sp_level2_checkbox, sp_level2_threshold=sp_level2_threshold, sp_level2_ge=sp_level2_ge,
        exclude_mt=exclude_mt, mt_threshold=mt_threshold,
        exclude_power_prod=exclude_power_prod, power_prod_threshold=power_prod_threshold, power_prod_ge=power_prod_ge,
        exclude_capacity=exclude_capacity, capacity_threshold=capacity_threshold, capacity_ge=capacity_ge,
        expansion_exclude=[e.strip() for e in expansion_exclude if e.strip()]
    )

    # 🔹 FIX: Added .copy() to prevent SettingWithCopyWarning and ensure Excluded/Exclusion Reasons
    #         columns are correctly written back to the DataFrame slices. Without .copy(), assignments
    #         to df["Excluded"] and df["Exclusion Reasons"] silently fail on DataFrame slices,
    #         causing all Urgewald-only (and other) rows to appear empty in the output. 🔹
    def apply(df):
        df = df.copy()  # ← FIX: prevents silent write failures on DataFrame slices
        if df.empty:
            return df.assign(Excluded=False, **{"Exclusion Reasons": ""})
        res = df.apply(lambda r: compute_exclusion(r, **params), axis=1, result_type="expand")
        df["Excluded"], df["Exclusion Reasons"] = res["Excluded"], res["Exclusion Reasons"]
        return df

    sp_merged = apply(sp_merged)
    sp_only = apply(sp_only)
    ur_unmerged = apply(ur_unmerged)

    # 🔹 Combines all companies that were flagged as "Excluded" across all sources into one master table. 🔹
    excluded_final = pd.concat([sp_merged[sp_merged.Excluded], sp_only[sp_only.Excluded], ur_unmerged[ur_unmerged.Excluded]])
    # 🔹 These are the retained companies (not excluded) from each source. 🔹
    retained_merged = sp_merged[~sp_merged.Excluded]
    sp_retained = sp_only[~sp_only.Excluded]
    ur_retained = ur_unmerged[~ur_unmerged.Excluded]

    final_cols = [
        "SP_ENTITY_NAME", "SP_ENTITY_ID", "SP_COMPANY_ID", "SP_ISIN", "SP_LEI",
        "Coal Industry Sector", "Company", ">10MT / >5GW",
        "Installed Coal Power Capacity (MW)",
        "Coal Share of Power Production", "Coal Share of Revenue", "expansion",
        "Generation (Thermal Coal)", "Thermal Coal Mining",
        "BB Ticker", "ISIN equity", "LEI", "Excluded", "Exclusion Reasons"
    ]
    
    # 🔹 FIX: Added .copy() to prevent SettingWithCopyWarning and ensure column assignments
    #         (adding missing columns, cleaning BB Ticker) are correctly applied. Without .copy(),
    #         modifications to DataFrame slices passed into this function are silently dropped,
    #         which caused Urgewald-only rows to show blank or missing data in the Excel output. 🔹
    def finalize(d):
        d = d.copy()  # ← FIX: prevents silent write failures on DataFrame slices
        for c in final_cols:
            if c not in d:
                d[c] = ""
        # strip the trailing " Equity " (and any whitespace before it) from BB tickers
        if "BB Ticker" in d:
            d["BB Ticker"] = (
                d["BB Ticker"]
                  .astype(str)
                  .str.replace(r"\s*Equity\s*$", "", regex=True)
                  .str.strip()
            )
        return d[final_cols]

    # 🔹 Output file creation 🔹
    excluded_final = finalize(excluded_final)
    retained_merged = finalize(retained_merged)
    sp_retained = finalize(sp_retained)
    ur_retained = finalize(ur_retained)
    
    buf = io.BytesIO()
     # 🔹 It generates an Excel file in memory with 4 neatly organized sheets containing the filtered results — ready for download in Streamlit. 🔹
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        excluded_final.to_excel(w, "Excluded Companies", index=False)
        retained_merged.to_excel(w, "Retained Companies", index=False)
        sp_retained.to_excel(w, "S&P Only", index=False)
        ur_retained.to_excel(w, "Urgewald Only", index=False)

    st.subheader("Results Summary")
    st.write(f"Excluded Companies: {len(excluded_final)}")
    st.write(f"Retained Companies (Merged & Retained): {len(retained_merged)}")
    st.write(f"S&P Only (Unmatched, Retained): {len(sp_retained)}")
    st.write(f"Urgewald Only (Unmatched, Retained): {len(ur_retained)}")

    st.download_button(
        label="Download Filtered Results",
        data=buf.getvalue(),
        file_name="Coal_Companies_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    main()
